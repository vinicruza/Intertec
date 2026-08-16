"""Gera a planilha de CMV dos pedidos da ORDEM DE PRODUÇÃO — julho e agosto/2026."""
import collections
import re
from decimal import Decimal, ROUND_HALF_UP

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from pedidos import ler
from resolver import EMBALAGEM_KIT, resolve

SAIDA = "CMV_Ordem_de_Producao_jul_ago_2026.xlsx"

AZUL = "1F3864"
CINZA = "F2F2F2"
VERDE = "E2EFDA"
AMARELO = "FFF2CC"
VERMELHO = "FCE4EC"

CORES_CONF = {"Alta": VERDE, "Média": AMARELO, "A confirmar": VERMELHO, "Sem CMV": VERMELHO}


def d2(v):
    if v is None:
        return None
    return float(Decimal(v).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def d6(v):
    if v is None:
        return None
    return float(Decimal(v).quantize(Decimal("0.000001"), rounding=ROUND_HALF_UP))


def cabecalho(ws, titulos, larguras):
    ws.append(titulos)
    for c in range(1, len(titulos) + 1):
        cel = ws.cell(row=1, column=c)
        cel.font = Font(bold=True, color="FFFFFF", size=10)
        cel.fill = PatternFill("solid", fgColor=AZUL)
        cel.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(c)].width = larguras[c - 1]
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"


def coleta():
    dados = []
    for aba, mes in [("jul26", "Julho/2026"), ("ago26", "Agosto/2026")]:
        for p in ler(aba):
            itens = []
            for it in p["itens"]:
                r = resolve(it["descricao"])
                q = it["qtd"]
                total = (q * r["cmv"]) if (q is not None and r["cmv"] is not None) else None
                conf = r["confianca"] if r["cmv"] is not None else "Sem CMV"
                itens.append({**it, "res": r, "total": total, "conf": conf})
            dados.append({**p, "mes": mes, "itens": itens})
    return dados


def aba_itens(wb, dados):
    ws = wb.create_sheet("Itens")
    cabecalho(ws,
              ["Mês", "Data", "Pág.", "Linha", "Cliente", "Qtd", "Descrição na ordem",
               "Produto no catálogo (Rentabilidade)", "CMV unitário (R$)",
               "CMV total (R$)", "Confiança", "Como foi lido / o que falta"],
              [11, 11, 6, 7, 26, 8, 42, 44, 14, 14, 12, 62])
    for p in dados:
        for it in p["itens"]:
            r = it["res"]
            produto = r["produto"] if r["produto"] else "— não identificado —"
            if r["tipo"] == "kit" and r["componentes"]:
                produto = "KIT montado (ver aba “Kits — composição”)"
            ws.append([p["mes"], p["data"], p["pagina"], it["linha"], p["cliente"],
                       float(it["qtd"]) if it["qtd"] is not None else None,
                       it["descricao"].strip(), produto,
                       d6(r["cmv"]), d2(it["total"]), it["conf"], r["obs"]])
            ws.cell(row=ws.max_row, column=11).fill = PatternFill(
                "solid", fgColor=CORES_CONF[it["conf"]])
    formata(ws, {9: "#,##0.000000", 10: "#,##0.00"}, data_col=2)
    return ws


def aba_pedidos(wb, dados):
    ws = wb.create_sheet("Pedidos")
    cabecalho(ws,
              ["Mês", "Data", "Pág.", "Cliente", "Itens", "CMV total do pedido (R$)",
               "Itens sem CMV", "Confiança do pedido", "Observações"],
              [11, 11, 6, 30, 7, 20, 13, 17, 70])
    for p in dados:
        total = sum((it["total"] for it in p["itens"] if it["total"] is not None), Decimal(0))
        sem = sum(1 for it in p["itens"] if it["total"] is None)
        confs = {it["conf"] for it in p["itens"]}
        conf = ("Sem CMV" if "Sem CMV" in confs and len(confs) == 1 else
                "A confirmar" if confs & {"Sem CMV", "A confirmar"} else
                "Média" if "Média" in confs else "Alta")
        obs = "; ".join(dict.fromkeys(n for n in p["notas"]))
        ws.append([p["mes"], p["data"], p["pagina"], p["cliente"] or "— sem cliente na planilha —",
                   len(p["itens"]), d2(total), sem, conf, obs])
        ws.cell(row=ws.max_row, column=8).fill = PatternFill("solid", fgColor=CORES_CONF[conf])
    formata(ws, {6: "#,##0.00"}, data_col=2)
    return ws


def aba_kits(wb, dados):
    ws = wb.create_sheet("Kits — composição")
    cabecalho(ws,
              ["Mês", "Data", "Cliente", "Linha", "Qtd de kits", "Descrição do kit",
               "Componente (como escrito)", "Produto no catálogo (não estéril)",
               "Qtd por kit", "CMV unit. (R$)", "CMV do componente por kit (R$)", "Observação"],
              [11, 11, 26, 7, 11, 44, 24, 42, 10, 13, 16, 55])
    for p in dados:
        for it in p["itens"]:
            r = it["res"]
            if r["tipo"] != "kit" or not r["componentes"]:
                continue
            for c in r["componentes"]:
                sub = (c["qtd"] * c["cmv"]) if c["cmv"] is not None else None
                ws.append([p["mes"], p["data"], p["cliente"], it["linha"],
                           float(it["qtd"]) if it["qtd"] else None, it["descricao"].strip(),
                           c["texto"], c["produto"] or "— não identificado —",
                           float(c["qtd"]), d6(c["cmv"]), d2(sub), c["obs"]])
            ws.append([p["mes"], p["data"], p["cliente"], it["linha"],
                       float(it["qtd"]) if it["qtd"] else None, it["descricao"].strip(),
                       "embalagem + esterilização do kit",
                       "padrão “Kit Aleatório” da Rentabilidade", 1.0,
                       d6(EMBALAGEM_KIT), d2(EMBALAGEM_KIT),
                       "caixa + envelope 30x40 + esterilização + etiqueta + gráfica; "
                       "o envelope e o rateio da caixa reais dependem do sistema da Intertech"])
            for c in range(1, 13):
                ws.cell(row=ws.max_row, column=c).font = Font(italic=True)
    formata(ws, {10: "#,##0.000000", 11: "#,##0.00"}, data_col=2)
    return ws


def formata(ws, formatos, data_col=None):
    fina = Side(style="thin", color="D9D9D9")
    for row in ws.iter_rows(min_row=2):
        for cel in row:
            if cel.font.size != 10:
                cel.font = Font(size=10, bold=cel.font.bold, italic=cel.font.italic)
            cel.border = Border(bottom=fina)
            cel.alignment = Alignment(vertical="top", wrap_text=cel.column in (7, 8, 12))
            if cel.column in formatos:
                cel.number_format = formatos[cel.column]
            if data_col and cel.column == data_col:
                cel.number_format = "dd/mm/yyyy"


def aba_resumo(wb, dados):
    ws = wb.create_sheet("Resumo", 0)
    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 60

    def titulo(txt):
        ws.append([txt])
        c = ws.cell(row=ws.max_row, column=1)
        c.font = Font(bold=True, size=12, color=AZUL)

    def linha(*vals, negrito=False, fmt=None):
        ws.append(list(vals))
        for i in range(1, len(vals) + 1):
            c = ws.cell(row=ws.max_row, column=i)
            c.font = Font(bold=negrito, size=10)
            if fmt and i > 1 and isinstance(vals[i - 1], float):
                c.number_format = fmt

    ws.append(["CMV dos pedidos — Ordem de Produção, julho e agosto/2026"])
    ws.cell(row=1, column=1).font = Font(bold=True, size=15, color=AZUL)
    linha("Base de custo: aba “Alocação Despesa” da planilha Rentabilidade 2026 "
          "(CMV unitário por produto).")
    linha("Regra de cálculo: Calculations.md §3 (CMV do produto) e §4.1 (CMV de kit "
          "= Σ componentes não estéreis + embalagem/esterilização do kit).")
    linha("")

    por_mes = collections.defaultdict(lambda: collections.defaultdict(Decimal))
    qtd_mes = collections.Counter()
    itens_mes = collections.Counter()
    sem_mes = collections.Counter()
    for p in dados:
        qtd_mes[p["mes"]] += 1
        for it in p["itens"]:
            itens_mes[p["mes"]] += 1
            if it["total"] is None:
                sem_mes[p["mes"]] += 1
            else:
                por_mes[p["mes"]][it["conf"]] += it["total"]

    titulo("1. Total por mês")
    linha("Mês", "Pedidos", "Itens", "CMV total (R$)", negrito=True)
    geral = Decimal(0)
    for mes in ["Julho/2026", "Agosto/2026"]:
        t = sum(por_mes[mes].values(), Decimal(0))
        geral += t
        linha(mes, float(qtd_mes[mes]), float(itens_mes[mes]), d2(t), fmt="#,##0.00")
    linha("TOTAL", float(sum(qtd_mes.values())), float(sum(itens_mes.values())),
          d2(geral), negrito=True, fmt="#,##0.00")
    linha("")

    titulo("2. Quanto do total é seguro")
    linha("Confiança", "CMV (R$)", "% do total", "Itens", negrito=True)
    cnt = collections.Counter(it["conf"] for p in dados for it in p["itens"])
    for conf, expl in [("Alta", "abreviação bate direto com um produto do catálogo"),
                       ("Média", "faltava a medida ou a gramatura; usei o padrão documentado"),
                       ("A confirmar", "leitura provável, precisa do seu OK")]:
        v = sum(por_mes[m][conf] for m in por_mes)
        linha(f"{conf} — {expl}", d2(v), float(round(v / geral * 100, 1)),
              float(cnt[conf]), fmt="#,##0.00")
    linha("Sem CMV — não deu para calcular", 0.0, 0.0, float(cnt["Sem CMV"]), fmt="#,##0.00")
    linha("")
    linha("Os itens “Sem CMV” são quase todos kits escritos só como “kit”, “kit 1”, "
          "“kit padrão”: a composição desses kits está no sistema da Intertech, não na planilha.")
    linha("")

    titulo("3. Sensibilidade — a gramatura que a ordem não diz")
    linha("Quando a ordem não escreve GR30 nem GR40, assumi GR40 (é a gramatura de maior "
          "produção estimada na Rentabilidade).")
    linha("Cenário", "CMV total (R$)", "Diferença", negrito=True)
    linha("Assumindo GR40 (o que está nesta planilha)", d2(geral), "—", fmt="#,##0.00")
    linha("Assumindo GR30", d2(TOTAL_GR30), d2(TOTAL_GR30 - geral), fmt="#,##0.00")
    linha("A escolha mexe em 1,2% do total — não é o ponto crítico.")
    linha("")

    titulo("4. Como ler as outras abas")
    linha("Itens", "uma linha por item da ordem, com o produto que encontrei e o CMV")
    linha("Pedidos", "uma linha por pedido (bloco da ordem), com o CMV somado")
    linha("Kits — composição", "cada kit aberto componente a componente")
    linha("Dicionário", "todas as abreviações que consegui entender")
    linha("Pendências", "o que precisa da sua confirmação, do maior impacto para o menor")
    linha("")
    titulo("5. As três perguntas que mais mudam o número")
    linha("1) Os kits fixos (“kit”, “kit 1”, “kit padrão”, “kit vet 4”, “kit uni 6”): "
          "qual a composição de cada um? É o maior buraco — sem isso não dá para fechar "
          f"{cnt['Sem CMV']} itens.")
    linha("2) O que é “BF” em “catarata BF” / “cataratinha 80 BF”? Não achei no catálogo "
          "nem na nomenclatura do sistema.")
    linha("3) “av sm” é Avental Sem Manga SMS ou Avental TNT Sem Manga? E “av marinho” "
          "sai de qual avental do catálogo?")
    for r in range(1, ws.max_row + 1):
        ws.cell(row=r, column=1).alignment = Alignment(vertical="top", wrap_text=False)
    return ws


def total_com_gramatura(g):
    import catalogo
    anterior = catalogo.GRAM_PADRAO
    catalogo.GRAM_PADRAO = g
    t = Decimal(0)
    for aba in ("jul26", "ago26"):
        for p in ler(aba):
            for it in p["itens"]:
                r = resolve(it["descricao"])
                if r["cmv"] is not None and it["qtd"] is not None:
                    t += it["qtd"] * r["cmv"]
    catalogo.GRAM_PADRAO = anterior
    return t


TOTAL_GR30 = Decimal(0)


def main():
    global TOTAL_GR30
    TOTAL_GR30 = total_com_gramatura("30")
    dados = coleta()
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    aba_itens(wb, dados)
    aba_pedidos(wb, dados)
    aba_kits(wb, dados)
    aba_dicionario(wb, dados)
    aba_pendencias(wb, dados)
    aba_resumo(wb, dados)
    wb.save(SAIDA)
    print("gerado", SAIDA)




DICIONARIO = [
    ("Famílias de produto", "", "", ""),
    ("catarata", "Campo Catarata", "Campo Catarata 1,00 x 1,20 (medida padrão)", "Alta"),
    ("cataratinha", "Campo Catarata pequeno", "Campo Catarata 0,60 x 0,60", "Alta"),
    ("cataratinha 80", "Campo Catarata 0,80 x 0,80", "Campo Catarata 0,80 x 0,80", "Alta"),
    ("sem fen", "campo SEM fenestra", "Campo Simples", "Alta"),
    ("com fen", "campo COM fenestra", "Campo Com Fenestra", "Alta"),
    ("com ades", "campo com adesivo", "Campo Com Adesivo", "Alta"),
    ("mesa", "campo de mesa", "Campo de Mesa", "Alta"),
    ("mayo", "campo de Mayo", "Campo de Mayo", "Alta"),
    ("ocl", "oclusor", "Oclusor", "Alta"),
    ("bino / mono", "campo Lasik binocular / monocular", "Campo Lasik Binocular / Monocular", "Alta"),
    ("bino com 2 bags", "Lasik binocular com 2 bags", "Campo Lasik Binocular 1,00 X 1,20 2 Bags", "Alta"),
    ("drape G / drape P", "steri drape grande / pequeno", "Steri Drape Grande / Pequeno", "Alta"),
    ("lat / sup / inf", "campo lateral / superior / inferior", "Campo Lateral / Superior / Inferior", "Alta"),
    ("saco", "saco", "Saco 0,40 x 0,60 (medida padrão)", "Média"),
    ("av", "avental", "Avental", "Alta"),
    ("cj / Cj", "conjunto (calça + blusa)", "Conjunto P/M ou G/GG", "Alta"),
    ("comp G / comp P", "compressa grande / pequena", "Compressa G / Compressa P", "Alta"),
    ("comp X com 5 / com 10", "pacote de compressas", "Compressa X Pacote 5 / Pacote 10", "Alta"),
    ("Toalha / T", "toalha de mão", "Compressa Wiper (na NF é “Toalha de Mão”)", "Alta"),
    ("perneira / bota", "perneira / bota", "Perneira / Bota", "Alta"),
    ("mangueira / refletor", "acessório de kit odonto", "não existe na Rentabilidade", "Sem CMV"),
    ("", "", "", ""),
    ("Medidas", "", "", ""),
    ("80, 60, 50, 40", "centímetros → campo quadrado", "0,80 x 0,80; 0,60 x 0,60; …", "Alta"),
    ("1 / 1,5 / 2", "metros → campo quadrado", "1,00 x 1,00; 1,50 x 1,50; 2,00 x 2,00", "Média"),
    ("1x1,4", "1,00 m x 1,40 m", "… 1,00 x 1,40", "Alta"),
    ("70x1", "0,70 m x 1,00 m", "… 0,70 x 1,00", "Alta"),
    ("1,2x2,00 / 1,2x200", "1,20 m x 2,00 m", "… 2,00 x 1,20 (o catálogo escreve o maior lado antes)", "Alta"),
    ("sem medida", "usa a medida mais vendida da família", "ver aba Pendências", "Média"),
    ("", "", "", ""),
    ("Modificadores", "", "", ""),
    ("não / nao / nãp", "NÃO estéril", "sufixo “Não Estéril”", "Alta"),
    ("(sem “não”)", "estéril — é o padrão da planilha", "produto estéril", "Alta"),
    ("CH", "origem China", "sufixo “China” (só existe em Campo Catarata GR30)", "Alta"),
    ("GR30 / GR40 / GR 40", "gramatura do TNT/SMS", "sufixo GR30 / GR40", "Alta"),
    ("TNT / SMS / lam", "tecido: TNT, SMS ou laminado", "variante do produto", "Alta"),
    ("bco / marinho / azul bebê", "cor do tecido", "o catálogo não separa por cor", "A confirmar"),
    ("com T", "com Tag (na NF a Tag vira “Toalha”)", "… com Tag", "Alta"),
    ("com P / com G", "com compressa P ou G", "… com Compressa (o catálogo não separa o tamanho)", "A confirmar"),
    ("ML", "manga longa", "Avental TNT 30g/40g ML", "Alta"),
    ("SM", "sem manga", "Avental Sem Manga SMS ou Avental TNT Sem Manga", "A confirmar"),
    ("GO", "ginecológico", "Avental Gineco", "Média"),
    ("especial", "tamanho especial", "… Tam Especial", "Média"),
    ("fen 10 / fen 12 / tape 4", "fenestra e fita extras", "só existe no catálogo em alguns produtos", "A confirmar"),
    ("embrulho", "embalado em vez de envelopado", "não muda de produto no catálogo", "A confirmar"),
    ("", "", "", ""),
    ("Kits", "", "", ""),
    ("kit 2M, 2T, 1 sem fen, 1 mesa 1x1,4", "composição escrita item a item",
     "soma dos componentes NÃO estéreis + embalagem/esterilização do kit", "Média"),
    ("2M / 2G / 2GG", "quantidade + tamanho de avental", "Avental / Avental G / Avental GG", "Alta"),
    ("2T", "2 toalhas de mão", "Compressa Wiper", "Alta"),
    ("kit vet 1..4", "kit veterinário do cliente", "Kit Veterinário (composição real no sistema)", "A confirmar"),
    ("kit uni 1/2/6", "kit universal do cliente", "Kit Universal Com Avental", "A confirmar"),
    ("kit I 40 / kit I gr 40", "kit odonto implante", "Kit Odonto Implante", "A confirmar"),
    ("kit / kit 1 / kit padrão", "kit fixo do cliente", "composição só existe no sistema da Intertech", "Sem CMV"),
]


def aba_dicionario(wb, dados):
    usos = collections.Counter()
    for p in dados:
        for it in p["itens"]:
            usos[it["descricao"].strip().lower()] += 1
    ws = wb.create_sheet("Dicionário")
    cabecalho(ws, ["Como aparece na ordem", "O que quer dizer",
                   "Produto no catálogo da Rentabilidade", "Segurança da leitura"],
              [34, 40, 58, 18])
    for a, b, c, d in DICIONARIO:
        ws.append([a, b, c, d])
        if b == "" and a:
            for i in range(1, 5):
                cel = ws.cell(row=ws.max_row, column=i)
                cel.font = Font(bold=True, color=AZUL, size=11)
                cel.fill = PatternFill("solid", fgColor=CINZA)
        elif d in CORES_CONF:
            ws.cell(row=ws.max_row, column=4).fill = PatternFill("solid", fgColor=CORES_CONF[d])
    formata(ws, {})
    return ws


def aba_pendencias(wb, dados):
    agrupado = collections.defaultdict(lambda: {"itens": 0, "qtd": Decimal(0),
                                                "cmv": Decimal(0), "exemplos": []})
    for p in dados:
        for it in p["itens"]:
            r = it["res"]
            if it["conf"] == "Alta":
                continue
            for msg in [m.strip() for m in (r["obs"] or "").split(";") if m.strip()]:
                g = agrupado[msg]
                g["itens"] += 1
                g["qtd"] += it["qtd"] or Decimal(0)
                g["cmv"] += it["total"] or Decimal(0)
                if len(g["exemplos"]) < 3 and it["descricao"].strip() not in g["exemplos"]:
                    g["exemplos"].append(it["descricao"].strip())
    ws = wb.create_sheet("Pendências")
    cabecalho(ws, ["O que precisa da sua confirmação", "Itens", "Peças",
                   "CMV envolvido (R$)", "Exemplos na ordem"], [78, 8, 10, 18, 52])
    for chave, g in sorted(agrupado.items(), key=lambda kv: -kv[1]["cmv"]):
        ws.append([chave, float(g["itens"]), float(g["qtd"]), d2(g["cmv"]),
                   " | ".join(g["exemplos"])])
    formata(ws, {4: "#,##0.00"})
    return ws


if __name__ == "__main__":
    main()
