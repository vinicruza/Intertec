"""Gera a planilha DE-PARA das abreviações da Ordem de Produção, para o cliente conferir.

Uma linha por abreviação distinta usada em julho e agosto/2026, com o que foi
entendido, o produto do catálogo e duas colunas em branco para a resposta.
"""
import collections
from decimal import Decimal, ROUND_HALF_UP

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from pedidos import ler
from resolver import EMBALAGEM_KIT, resolve

SAIDA = "De-Para_Abreviacoes_Ordem_de_Producao.xlsx"

AZUL = "1F3864"
CINZA = "F2F2F2"
VERDE = "E2EFDA"
AMARELO = "FFF2CC"
VERMELHO = "FCE4EC"
RESPOSTA = "DDEBF7"


def d2(v):
    return None if v is None else float(Decimal(v).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def d6(v):
    return None if v is None else float(Decimal(v).quantize(Decimal("0.000001"),
                                                            rounding=ROUND_HALF_UP))


def cabecalho(ws, titulos, larguras, cols_resposta=()):
    ws.append(titulos)
    for c in range(1, len(titulos) + 1):
        cel = ws.cell(row=1, column=c)
        cel.font = Font(bold=True, color="FFFFFF", size=10)
        cel.fill = PatternFill("solid", fgColor="7B3F00" if c in cols_resposta else AZUL)
        cel.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(c)].width = larguras[c - 1]
    ws.row_dimensions[1].height = 34
    ws.freeze_panes = "A2"


def formata(ws, quebra=(), cols_resposta=(), fmt=None):
    fina = Side(style="thin", color="D9D9D9")
    for row in ws.iter_rows(min_row=2):
        for cel in row:
            if cel.font.size != 10:
                cel.font = Font(size=10, bold=cel.font.bold)
            cel.border = Border(bottom=fina)
            cel.alignment = Alignment(vertical="top", wrap_text=cel.column in quebra)
            if cel.column in cols_resposta:
                cel.fill = PatternFill("solid", fgColor=RESPOSTA)
            if fmt and cel.column in fmt:
                cel.number_format = fmt[cel.column]


def valida_sim_nao(ws, coluna, primeira=2):
    dv = DataValidation(type="list", formula1='"Sim,Não"', allow_blank=True)
    dv.prompt = "Marque Sim se a leitura está certa, ou Não e escreva o certo ao lado."
    dv.promptTitle = "Confere?"
    ws.add_data_validation(dv)
    dv.add(f"{coluna}{primeira}:{coluna}{max(ws.max_row, primeira)}")


def coleta():
    """Agrupa por abreviação distinta (texto exato como escrito na ordem)."""
    grupos = collections.defaultdict(lambda: {
        "vezes": 0, "pecas": Decimal(0), "clientes": [], "meses": set()})
    for aba, mes in [("jul26", "Julho/2026"), ("ago26", "Agosto/2026")]:
        for p in ler(aba):
            for it in p["itens"]:
                chave = it["descricao"].strip()
                g = grupos[chave]
                g["vezes"] += 1
                g["pecas"] += it["qtd"] or Decimal(0)
                g["meses"].add(mes)
                if p["cliente"] and p["cliente"] not in g["clientes"] and len(g["clientes"]) < 4:
                    g["clientes"].append(p["cliente"])
    for chave, g in grupos.items():
        g["res"] = resolve(chave)
        g["cmv_periodo"] = (g["pecas"] * g["res"]["cmv"]) if g["res"]["cmv"] is not None else None
    return grupos


def reais(v):
    """Formata em padrão brasileiro: 547.431,73"""
    return f"{d2(v):,.2f}".replace(",", "@").replace(".", ",").replace("@", ".")


def composicao_texto(res):
    if res["tipo"] != "kit" or not res["componentes"]:
        return ""
    comps = res["componentes"]
    if len(comps) == 1 and str(comps[0]["produto"]).startswith("Kit"):
        return f"{comps[0]['produto']} (kit pronto do catálogo, usado como referência)"
    partes = [f"{int(c['qtd'])}× {c['produto'] or '?'}" for c in comps]
    partes.append("+ embalagem/esterilização do kit")
    return "; ".join(partes)


# --------------------------------------------------------------- abas
def aba_instrucoes(wb, grupos):
    ws = wb.create_sheet("Como preencher")
    ws.column_dimensions["A"].width = 118
    def p(txt, negrito=False, tam=10, cor="000000"):
        ws.append([txt])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=negrito, size=tam, color=cor)
        ws.cell(row=ws.max_row, column=1).alignment = Alignment(wrap_text=True, vertical="top")

    p("De-Para das abreviações da Ordem de Produção", True, 15, AZUL)
    p("Julho e agosto/2026 — para conferência da Intertech", False, 11)
    p("")
    p("Por que esta planilha existe", True, 12, AZUL)
    p("A Ordem de Produção é escrita em abreviação: “200 ocl”, “catarata CH não”, "
      "“kit 2M, 2T, 1 sem fen”. Para calcular o CMV de cada pedido é preciso saber qual "
      "produto do catálogo cada apelido representa. Aqui está tudo o que eu entendi — e o "
      "que eu não consegui.")
    p("")
    p("O que fazer em cada aba", True, 12, AZUL)
    p("1. Dúvidas — comece por aqui. São as perguntas que mudam o valor do CMV. "
      "Poucas linhas, e cada resposta destrava dezenas de itens.")
    p("2. Kits sem composição — os kits que aparecem só como “kit”, “kit 1”, “kit padrão”. "
      "Escreva a composição de cada um (ex.: 2 aventais M, 2 toalhas, 1 campo simples 1x1,2).")
    p("3. De-para completo — todas as abreviações usadas nos dois meses, da mais usada para a "
      "menos. Confira por cima; o que estiver errado, marque “Não” e escreva o produto certo.")
    p("4. Regras de leitura — as regras gerais que apliquei (o que é medida, o que é "
      "gramatura, o que quer dizer “não”). Confirme as regras e o resto se resolve sozinho.")
    p("")
    p("As colunas em azul claro são suas", True, 12, AZUL)
    p("“Confere?” tem uma listinha de Sim/Não. Se marcar Não, escreva ao lado qual é o "
      "produto certo (pode ser o nome como está no catálogo da Rentabilidade, ou como vocês "
      "chamam — eu faço a ponte).")
    p("")
    p("Como está hoje", True, 12, AZUL)
    total = sum(g["cmv_periodo"] or Decimal(0) for g in grupos.values())
    n_conf = sum(1 for g in grupos.values() if g["res"]["confianca"] == "A confirmar"
                 or g["res"]["cmv"] is None)
    p(f"São {len(grupos)} abreviações diferentes nos dois meses, somando "
      f"R$ {reais(total)} de CMV calculado. Dessas, {n_conf} estão marcadas como "
      "“a confirmar” — é exatamente o que esta planilha quer resolver.")
    return ws


PERGUNTAS = [
    ("BF", "“catarata BF”, “cataratinha 80 BF”, “Catarata BF”",
     "Não achei. Calculei como se o BF não existisse (Campo Catarata comum).",
     "O que é “BF”? É outro produto, outro acabamento, ou um cliente?"),
    ("av sm", "“av sm”, “av sm bco”",
     "Usei Avental Sem Manga SMS.",
     "É o Avental Sem Manga SMS ou o Avental TNT Sem Manga? Os dois existem no catálogo, "
     "com custo diferente."),
    ("av marinho", "“av marinho”, “av marinho especial”, “Av Azul bebe GO”",
     "O catálogo não tem cor. Usei o Avental comum (e Avental Gineco no “Azul bebe GO”).",
     "Avental marinho sai de qual avental do catálogo? A cor muda o custo do tecido?"),
    ("com P / com G", "“Av M com P”, “Av G com G”, “Av GG com P”",
     "Entendi como “com Compressa” — o catálogo não separa o tamanho da compressa.",
     "O P e o G aí são o tamanho da compressa? Se o custo for diferente, precisamos de "
     "dois produtos no catálogo."),
    ("com B", "“Cj M com B”, “Cj G com B”, “Cg GG com B”",
     "Não achei. Conjunto no catálogo é calça + blusa, sem variante “com B”.",
     "O que entra a mais no conjunto “com B”?"),
    ("inga / inv", "“drape G inga”, “drape P inv”",
     "Calculei como Steri Drape Grande / Pequeno comum.",
     "O que quer dizer “inga” e “inv” no drape?"),
    ("sem fen 1 / mesa 1", "“sem fen 1”, “mesa 1”, “com fen 1”",
     "Li número sozinho como campo quadrado: “1” = 1,00 x 1,00.",
     "Confirma? Ou “sem fen 1” quer dizer 1,00 x 1,20 (o mais vendido)?"),
    ("medida em branco", "“catarata”, “mesa”, “sem fen”, “saco”",
     "Usei a medida mais vendida da família: catarata 1,00 x 1,20; mesa 1,00 x 1,40; "
     "sem fen 1,00 x 1,20; saco 0,40 x 0,60.",
     "Confirma esses padrões? São 130 itens, o maior bloco de suposição."),
    ("gramatura em branco", "“catarata”, “sem fen 80”, “com fen 60”…",
     "Quando não diz GR30 nem GR40, assumi GR40.",
     "Confirma GR40 como padrão? Em GR30 o CMV total cai 1,2%."),
    ("medidas que não existem", "“sem fen 1x1,5”, “sem fen 1,4x2,00”, “sem fen 1,6x2,00”, "
     "“com fen 20”, “comp G com 2”",
     "Ficaram sem CMV — o catálogo da Rentabilidade não tem esses tamanhos.",
     "Falta cadastrar esses produtos, ou é apelido de algum que já existe?"),
    ("fenestra e tape extras", "“com fen 50, fen 6”, “com fen 1x1,2, fen 10, tape 4”, "
     "“sem fen 60, tape 60”",
     "Usei o campo sem o extra — o catálogo só tem alguns tamanhos com fenestra/tape.",
     "O extra muda o custo? Se muda, viram produtos próprios."),
    ("embrulho", "“sem fen 1,4x2,00 embrulho”, “mesa 1x1,4 embrulho”",
     "Não mudei nada no custo.",
     "“Embrulho” troca a embalagem (sem envelope)? Se troca, o CMV muda."),
    ("mangueira / refletor", "dentro de kits odonto",
     "Ficaram sem custo — não têm ficha na Rentabilidade.",
     "Vocês compram esses itens? Qual o custo unitário?"),
    ("Av ML com elast", "“Av ML com elast”",
     "Ficou sem CMV — só existe o insumo “costureira avental com elástico”, não o produto.",
     "Qual avental do catálogo corresponde?"),
    ("embalagem do kit", "todos os kits montados",
     "Usei o padrão “Kit Aleatório” da própria Rentabilidade: R$ 2,026695 "
     "(caixa + envelope 30x40 + esterilização + etiqueta + gráfica).",
     "Qual envelope e quantos kits por caixa de esterilização, por tipo de kit? "
     "É o que falta para o CMV do kit ficar 100%."),
]


def aba_duvidas(wb, grupos):
    ws = wb.create_sheet("1. Dúvidas")
    cabecalho(ws, ["Assunto", "Como aparece na ordem", "O que eu fiz",
                   "O que preciso saber", "Itens", "CMV envolvido (R$)",
                   "Sua resposta"],
              [22, 40, 46, 52, 8, 16, 46], cols_resposta=(7,))
    # impacto por assunto, medido pelas abreviações que casam com os exemplos
    for assunto, exemplos, feito, pergunta in PERGUNTAS:
        itens, cmv = impacto(grupos, assunto)
        ws.append([assunto, exemplos, feito, pergunta,
                   float(itens) if itens else None, d2(cmv) if cmv else None, ""])
    formata(ws, quebra=(2, 3, 4, 7), cols_resposta=(7,), fmt={6: "#,##0.00"})
    return ws


# Cada assunto é medido ou pelo texto da abreviação (t) ou pela observação que o
# resolvedor gravou (o) — a observação é mais precisa quando a regra é interna.
CHAVES_IMPACTO = {
    "BF": lambda t, o: "entendi “bf”" in o,
    "av sm": lambda t, o: "Avental Sem Manga SMS —" in o,
    "av marinho": lambda t, o: "cor “" in o,
    "com P / com G": lambda t, o: "com Compressa” —" in o,
    "com B": lambda t, o: "com B” não tem correspondente" in o,
    "inga / inv": lambda t, o: "inga" in t or " inv" in t,
    "sem fen 1 / mesa 1": lambda t, o: t in ("sem fen 1", "mesa 1", "com fen 1",
                                             "sem fen 1 não", "sem fen 1 nao"),
    "medida em branco": lambda t, o: "medida não informada" in o,
    "gramatura em branco": lambda t, o: "gramatura assumida" in o,
    "medidas que não existem": lambda t, o: "não existe no catálogo" in o
                                            or "não existe no catálogo (só Pacote" in o,
    "fenestra e tape extras": lambda t, o: "não entendi “fen" in o or "não entendi “tape" in o
                                           or "tape" in o and "não entendi" in o,
    "embrulho": lambda t, o: "embrulho" in t,
    "mangueira / refletor": lambda t, o: "mangueira" in t or "magueira" in t or "refletor" in t,
    "Av ML com elast": lambda t, o: "elast" in t,
    "embalagem do kit": lambda t, o: "Kit Aleatório" in o,
}


def impacto(grupos, assunto):
    teste = CHAVES_IMPACTO.get(assunto)
    if not teste:
        return 0, Decimal(0)
    itens = 0
    cmv = Decimal(0)
    for chave, g in grupos.items():
        if teste(chave.lower(), g["res"]["obs"] or ""):
            itens += g["vezes"]
            cmv += g["cmv_periodo"] or Decimal(0)
    return itens, cmv


def aba_kits(wb, grupos):
    ws = wb.create_sheet("2. Kits sem composição")
    cabecalho(ws, ["Como aparece na ordem", "Vezes", "Peças", "Clientes que pediram",
                   "O que eu usei no cálculo", "Composição real (escreva aqui)"],
              [40, 8, 10, 44, 46, 60], cols_resposta=(6,))
    linhas = []
    for chave, g in grupos.items():
        r = g["res"]
        if r["tipo"] != "kit":
            continue
        if r["cmv"] is None or (r["confianca"] == "A confirmar" and not r["componentes"]):
            usado = "nada — ficou sem CMV"
        elif r["componentes"] and len(r["componentes"]) == 1 and \
                r["componentes"][0]["produto"] and "Kit" in str(r["componentes"][0]["produto"]):
            usado = f"{r['componentes'][0]['produto']} (kit genérico do catálogo)"
        else:
            continue
        linhas.append((g["pecas"], [chave, float(g["vezes"]), float(g["pecas"]),
                                    ", ".join(g["clientes"]), usado, ""]))
    for _, l in sorted(linhas, key=lambda x: -x[0]):
        ws.append(l)
    formata(ws, quebra=(1, 4, 5, 6), cols_resposta=(6,))
    return ws


def aba_depara(wb, grupos):
    ws = wb.create_sheet("3. De-para completo")
    cabecalho(ws, ["Como aparece na ordem", "Vezes", "Peças", "Meses",
                   "O que eu entendi", "Produto no catálogo da Rentabilidade",
                   "CMV unit. (R$)", "CMV no período (R$)", "Situação",
                   "Confere?", "Se não, qual é o produto certo?"],
              [40, 7, 10, 20, 44, 46, 13, 15, 14, 11, 46], cols_resposta=(10, 11))
    linhas = []
    for chave, g in grupos.items():
        r = g["res"]
        if r["cmv"] is None:
            situacao = "Sem CMV"
        elif r["confianca"] == "Alta":
            situacao = "Certo"
        elif r["confianca"] == "Média":
            situacao = "Padrão assumido"
        else:
            situacao = "A confirmar"
        entendi = r["obs"] or "leitura direta"
        produto = r["produto"] or "— não identificado —"
        if r["tipo"] == "kit" and r["componentes"]:
            produto = composicao_texto(r)
        linhas.append((0 if situacao == "Certo" else 1, -(g["cmv_periodo"] or Decimal(0)),
                       [chave, float(g["vezes"]), float(g["pecas"]),
                        " e ".join(sorted(g["meses"])), entendi, produto,
                        d6(r["cmv"]), d2(g["cmv_periodo"]), situacao, "", ""]))
    for _, _, l in sorted(linhas, key=lambda x: (-x[0], x[1])):
        ws.append(l)
        cor = {"Certo": VERDE, "Padrão assumido": AMARELO,
               "A confirmar": VERMELHO, "Sem CMV": VERMELHO}[l[8]]
        ws.cell(row=ws.max_row, column=9).fill = PatternFill("solid", fgColor=cor)
    formata(ws, quebra=(1, 5, 6, 11), cols_resposta=(10, 11),
            fmt={7: "#,##0.000000", 8: "#,##0.00"})
    valida_sim_nao(ws, "J")
    return ws


REGRAS = [
    ("Famílias", "", "", ""),
    ("catarata", "Campo Catarata", "medida padrão 1,00 x 1,20", "Sim"),
    ("cataratinha", "Campo Catarata pequeno", "medida padrão 0,60 x 0,60", "Sim"),
    ("sem fen", "campo SEM fenestra", "família Campo Simples do catálogo", "Sim"),
    ("com fen", "campo COM fenestra", "família Campo Com Fenestra", "Sim"),
    ("com ades", "campo com adesivo", "família Campo Com Adesivo", "Sim"),
    ("mesa", "campo de mesa", "medida padrão 1,00 x 1,40", "Sim"),
    ("mayo / ocl / bino / mono", "Campo de Mayo, Oclusor, Lasik Binocular, Lasik Monocular",
     "", "Sim"),
    ("lat / sup / inf", "Campo Lateral, Superior, Inferior", "", "Sim"),
    ("av / cj / comp / drape", "Avental, Conjunto, Compressa, Steri Drape", "", "Sim"),
    ("T (sozinho, dentro de kit)", "toalha de mão", "= Compressa Wiper no catálogo", "Sim"),
    ("saco", "Saco", "medida padrão 0,40 x 0,60", "Sim"),
    ("", "", "", ""),
    ("Medidas", "", "", ""),
    ("um número só (80, 60, 1, 1,5)", "campo quadrado", "80 = 0,80 x 0,80; 1,5 = 1,50 x 1,50",
     "Sim"),
    ("dois números com x (1x1,4)", "os dois lados", "1x1,4 = 1,00 x 1,40; 70x1 = 0,70 x 1,00",
     "Sim"),
    ("número a partir de 20", "está em centímetro", "70 = 0,70 m", "Sim"),
    ("sem medida", "a medida mais vendida da família", "ver aba 1. Dúvidas", "Sim"),
    ("", "", "", ""),
    ("Acabamento", "", "", ""),
    ("não / nao", "NÃO estéril", "", "Sim"),
    ("sem a palavra “não”", "ESTÉRIL", "é o padrão da planilha", "Sim"),
    ("CH", "origem China", "só existe em Campo Catarata GR30", "Sim"),
    ("GR30 / GR40", "gramatura do tecido", "sem indicação, assumi GR40", "Sim"),
    ("TNT / SMS / lam", "tipo de tecido", "", "Sim"),
    ("com T", "com Tag", "na nota fiscal a Tag vira “Toalha”", "Sim"),
    ("ML / SM / GO", "manga longa / sem manga / ginecológico", "", "Sim"),
    ("", "", "", ""),
    ("Kits", "", "", ""),
    ("kit 2M, 2T, 1 sem fen, 1 mesa 1x1,4", "composição escrita item a item",
     "somo os componentes e acrescento a embalagem do kit", "Sim"),
    ("2M / 2G / 2GG dentro do kit", "quantidade + tamanho de avental",
     "M = o Avental padrão do catálogo", "Sim"),
    ("componente dentro do kit", "entra pelo custo NÃO ESTÉRIL",
     "porque a esterilização e o envelope são um só para o kit inteiro, "
     "não um por produto dentro dele", "Sim"),
]


def aba_regras(wb):
    ws = wb.create_sheet("4. Regras de leitura")
    cabecalho(ws, ["Como aparece na ordem", "O que eu entendi que é", "Detalhe",
                   "Confere?", "Se não, o certo é"],
              [36, 46, 52, 11, 46], cols_resposta=(4, 5))
    for a, b, c, _ in REGRAS:
        ws.append([a, b, c, "", ""])
        if b == "" and a:
            for i in range(1, 6):
                cel = ws.cell(row=ws.max_row, column=i)
                cel.font = Font(bold=True, color=AZUL, size=11)
                cel.fill = PatternFill("solid", fgColor=CINZA)
    formata(ws, quebra=(2, 3, 5), cols_resposta=(4, 5))
    for row in ws.iter_rows(min_row=2):
        if not row[1].value and row[0].value:
            for cel in row:
                cel.fill = PatternFill("solid", fgColor=CINZA)
    valida_sim_nao(ws, "D")
    return ws


def main():
    grupos = coleta()
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    aba_instrucoes(wb, grupos)
    aba_duvidas(wb, grupos)
    aba_kits(wb, grupos)
    aba_depara(wb, grupos)
    aba_regras(wb)
    wb.save(SAIDA)
    print("gerado", SAIDA, "—", len(grupos), "abreviações")


if __name__ == "__main__":
    main()
